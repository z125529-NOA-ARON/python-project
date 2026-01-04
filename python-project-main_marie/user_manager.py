import openpyxl
import os
from datetime import datetime


# --- Exceptions personnalisées pour la gestion des erreurs  ---
class UserManagerError(Exception):
    pass  


class EmptyFieldError(UserManagerError):
    pass  # Levée quand username/password sont vides


class InvalidCredentialsError(UserManagerError):
    pass  # Levée quand les identifiants sont invalides


class FileAccessError(UserManagerError):
    pass  # Levée quand le fichier Excel ne peut être lu/écrit


class User:
    def __init__(self, username, password, best_score=0, last_score=0, best_score_date=None):
        self.username = username
        self.password = password
        self.best_score = best_score
        self.last_score = last_score
        self.best_score_date = best_score_date

    def __lt__(self, other):
        # Permet de trier les utilisateurs par score (utilisé pour le classement)
        return self.best_score < other.best_score

    def __eq__(self, other):
        # Compare deux utilisateurs par leur nom d'utilisateur
        if isinstance(other, User):
            return self.username == other.username
        return False

    def __add__(self, value):
        # Permet User + int -> retourne un nouvel User avec score augmenté (non destructif)
        if isinstance(value, int):
            return User(self.username, self.password, self.best_score + value)
        return NotImplemented

    def __iadd__(self, value):
        # Permet += pour augmenter le score en place
        if isinstance(value, int):
            self.best_score += value
            return self
        return NotImplemented

    def __repr__(self):
        return f"User(username={self.username!r}, best_score={self.best_score})"



header = ["username", "password", "last_score", "best_score", "best_score_date"]

class UserManager:
    def __init__(self, filename="info_snake.xlsx"):
        self.filename = filename
        self.users = {}
        self.load_users()

    def load_users(self):
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(header)  # entête du fichier
            wb.save(self.filename)

        try:
            wb = openpyxl.load_workbook(self.filename)
        except Exception as e:
            # Propagation via exception spécifique pour faciliter les tests 
            raise FileAccessError(f"\033[31mImpossible de lire le fichier users\033[0m: {e}")

        sheet = wb.active
        self.users.clear()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            username, password, last_score, best_score, best_score_date = row

            self.users[username] = User(
                username=username,
                password=password,
                best_score=best_score or 0,
                last_score=last_score or 0,
                best_score_date=best_score_date
            )

    def save_to_history(self, history_filename="history_snake.xlsx"):
        HISTORY_HEADER = ["username", "best_score", "best_score_date"]

        if not os.path.exists(history_filename):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(HISTORY_HEADER)
            wb.save(history_filename)

        try:
            wb = openpyxl.load_workbook(history_filename)
            sheet = wb.active
        except Exception as e:
            raise FileAccessError(f"Impossible d'écrire l'historique: {e}")

        best_user = None
        best_score = -1

        for user in self.users.values():
            if user.best_score > best_score:
                best_score = user.best_score
                best_user = user

        if best_user is not None and best_user.best_score > 0:
            sheet.append([
                best_user.username,
                best_user.best_score,
                best_user.best_score_date
            ])

        wb.save(history_filename)



    def save_users(self):

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(header)

        for user in self.users.values():
            sheet.append([
                user.username,
                user.password,
                user.last_score,
                user.best_score,
                user.best_score_date
            ])

        try:
            wb.save(self.filename)
        except Exception as e:
            raise FileAccessError(f"Impossible d'écrire le fichier: {e}")


    def authenticate(self, username, password):
        # Validation minimale des champs
        if not username or not password:
            raise EmptyFieldError("\033[31mLe nom d'utilisateur et le mot de passe ne peuvent pas être vides\033[0m")

        if username not in self.users:
            # créer un nouvel utilisateur si inexistant
            self.users[username] = User(username, password, 0)
            self.save_users()
            return True

        # Comparaison du mot de passe existant
        user = self.users[username]
        if user.password == password:
            return True
        # Pour l'évaluation : lever une exception quand le mot de passe est incorrect
        raise InvalidCredentialsError("\033[31mPassword incorrect\033[0m")

    def update_best_score(self, username, score):
        if username not in self.users:
            raise InvalidCredentialsError("\033[31mUser unknown\033[0m")

        user = self.users[username]
        user.last_score = score
        if score > user.best_score:
            user.best_score = score
            user.best_score_date = datetime.now().strftime("%d/%m/%Y_%H:%M")
        
        self.save_users()

    def get_top_4(self):
        # tri décroissant
        # On retourne une structure compatible avec l'existant: (username, {"best_score": ...})
        sorted_users = sorted(
            list(self.users.items()),
            key=lambda x: x[1].best_score,
            reverse=True
        )

        result = []
        for username, user in sorted_users[:4]:
            result.append((username, {"best_score": user.best_score}))
        return result


    def reset_all(self):
        #Sauvegarder l'historique
        self.save_to_history()

        #Vider les utilisateurs
        self.users.clear()

        #Réécrire le fichier principal
        self.save_users()



