import openpyxl
import os


"""
user_manager.py
Provides a small user management system backed by an Excel file.

Responsibilities:
- define custom exceptions for caller-friendly error handling
- represent users with the `User` dataclass-like object
- load/save users from/to an Excel workbook using openpyxl
- authenticate users and update high scores

Errors are exposed via
specific exceptions and data is kept in-memory until saved.

Noa - Z125529
"""


class UserManagerError(Exception):
    """Base exception for all UserManager-related errors."""


class EmptyFieldError(UserManagerError):
    """Raised when a required field (username or password) is empty."""


class InvalidCredentialsError(UserManagerError):
    """Raised when authentication fails due to invalid credentials."""


class FileAccessError(UserManagerError):
    """Raised when the underlying Excel file cannot be read or written."""


class User:
    """Simple user object storing username, password and best_score.

    The class implements a few operator overloads to make it convenient
    to compare, display and increment scores.
    """

    def __init__(self, username, password, best_score=0):
        self.username = username
        self.password = password
        self.best_score = best_score

    def __lt__(self, other):
        # Allow sorting users by their best score (used for leaderboards)
        return self.best_score < other.best_score

    def __eq__(self, other):
        # Equality is defined by username only
        if isinstance(other, User):
            return self.username == other.username
        return False

    def __add__(self, value):
        # Return a new User with an increased score (non-destructive)
        if isinstance(value, int):
            return User(self.username, self.password, self.best_score + value)
        return NotImplemented

    def __iadd__(self, value):
        # In-place score increment (supports += int)
        if isinstance(value, int):
            self.best_score += value
            return self
        return NotImplemented

    def __repr__(self):
        return f"User(username={self.username!r}, best_score={self.best_score})"


class UserManager:
    """Manage users persisted in an Excel file.

    Usage:
        manager = UserManager('info_snake.xlsx')
        manager.authenticate('alice', 'password')
        manager.update_best_score('alice', 42)
    """

    def __init__(self, filename="info_snake.xlsx"):
        self.filename = filename
        self.users = {}
        self.load_users()

    def load_users(self):
        """Load users from the Excel file into memory.

        If the file does not exist, create it with a header row.
        Any file access error is re-raised as FileAccessError to make
        error handling predictable for callers and tests.
        """
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["username", "password", "best_score"])  # header row
            wb.save(self.filename)

        try:
            wb = openpyxl.load_workbook(self.filename)
        except Exception as e:
            # Wrap low-level exceptions into a domain-specific one
            raise FileAccessError(f"\033[31mUnable to read users file\033[0m: {e}")

        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            # Skip the header and construct User objects for each entry
            if row[0] != "username":
                best = row[2] if row[2] is not None else 0
                self.users[row[0]] = User(row[0], row[1], best)

    def save_users(self):
        """Persist the current `self.users` mapping into the Excel file.

        The file is rewritten entirely from the in-memory data to keep
        the implementation simple and deterministic.
        """
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["username", "password", "best_score"])  # header

        for username, user in self.users.items():
            # Each user is written as a row: username, password, best_score
            sheet.append([username, user.password, user.best_score])

        try:
            wb.save(self.filename)
        except Exception as e:
            raise FileAccessError(f"\033[31mUnable to write users file\033[0m: {e}")

    def authenticate(self, username, password):
        """Authenticate a user or create it if unknown.

        - Raises EmptyFieldError when username or password is empty.
        - Creates a new user (and saves) when username is not present.
        - Returns True when authentication succeeds.
        - Raises InvalidCredentialsError when the password is wrong.
        """
        if not username or not password:
            raise EmptyFieldError("\033[31mUsername and password cannot be empty\033[0m")

        if username not in self.users:
            # Create a new user if it does not exist yet
            self.users[username] = User(username, password, 0)
            self.save_users()
            return True

        # Compare the stored password for existing users
        user = self.users[username]
        if user.password == password:
            return True
        # Signal incorrect password via an exception
        raise InvalidCredentialsError("\033[31mPassword incorrect\033[0m")

    def update_best_score(self, username, score):
        """Update the best score for `username` if `score` is higher.

        Raises InvalidCredentialsError if the user does not exist.
        """
        if username not in self.users:
            raise InvalidCredentialsError("\033[31mUser unknown\033[0m")

        user = self.users[username]
        if score > user.best_score:
            user.best_score = score
            self.save_users()

    def get_top_4(self):
        """Return the top 4 users as a list of (username, {"best_score": ...}).

        The result matches the existing code's expected shape so callers
        can remain unchanged while benefiting from clearer implementation.
        """
        # Sort users by best_score descending
        sorted_users = sorted(
            list(self.users.items()),
            key=lambda x: x[1].best_score,
            reverse=True,
        )

        result = []
        for username, user in sorted_users[:4]:
            result.append((username, {"best_score": user.best_score}))
        return result

